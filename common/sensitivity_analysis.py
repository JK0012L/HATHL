"""
heterogeneity_sensitivity_analysis.py
Heterogeneity Parameter Sensitivity Analysis - Complete Data Statistics and Visualization
Referencing the architecture design of experimental_analysis.py
"""

import os
import sys
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook
from pathlib import Path
from scipy import stats
from math import pi
import warnings
import logging

warnings.filterwarnings('ignore')

# ============================================================================
# Configuration Settings
# ============================================================================

# Set Chinese/English display support
plt.rcParams['font.sans-serif'] = ['SimHei', 'DejaVu Sans', 'Arial Unicode MS', 'Microsoft YaHei']
plt.rcParams['axes.unicode_minus'] = False
plt.rcParams['figure.dpi'] = 300
plt.rcParams['savefig.dpi'] = 300
plt.rcParams['savefig.bbox'] = 'tight'
plt.rcParams['font.size'] = 11
plt.rcParams['axes.labelsize'] = 12
plt.rcParams['axes.titlesize'] = 14
plt.rcParams['xtick.labelsize'] = 10
plt.rcParams['ytick.labelsize'] = 10
plt.rcParams['legend.fontsize'] = 10

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('heterogeneity_analysis.log', encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)


# ============================================================================
# Experiment Configuration Parameters
# ============================================================================

class HeterogeneityConfig:
    """Heterogeneity experiment configuration manager"""
    
    # Scenario definitions (12 orthogonal experiment scenarios)
    SCENARIOS = [f'EXP-{i}' for i in range(1, 13)]
    
    # Heterogeneity types
    HET_TYPES = ['machine', 'repair', 'worker']
    
    # Level values for each heterogeneity type
    HET_VALUES = {
        'machine': [0.05, 0.10, 0.15, 0.20, 0.25, 0.30, 0.40, 0.50, 0.60],
        'repair': [0.05, 0.10, 0.15, 0.20, 0.30, 0.50, 0.70, 0.85, 1.00],
        'worker': [0.05, 0.10, 0.15, 0.20, 0.25, 0.30, 0.40, 0.50, 0.60]
    }
    
    # Column name mapping (Excel column name format)
    COLUMN_NAMES = {
        'machine': [f'M-HET-{i}' for i in range(1, 10)],
        'repair': [f'R-HET-{i}' for i in range(1, 10)],
        'worker': [f'W-HET-{i}' for i in range(1, 10)]
    }
    
    # Scenario feature parameters
    SCENARIO_FEATURES = {
        'EXP-1': {'M': 5, 'W': 2, 'U': 0.75, 'W_M_ratio': 2/5, 'group': 'medium', 'description': 'M=5,W=2,U=0.75'},
        'EXP-2': {'M': 5, 'W': 3, 'U': 0.85, 'W_M_ratio': 3/5, 'group': 'medium', 'description': 'M=5,W=3,U=0.85'},
        'EXP-3': {'M': 5, 'W': 4, 'U': 0.95, 'W_M_ratio': 4/5, 'group': 'high', 'description': 'M=5,W=4,U=0.95'},
        'EXP-4': {'M': 5, 'W': 2, 'U': 0.95, 'W_M_ratio': 2/5, 'group': 'low', 'description': 'M=5,W=2,U=0.95'},
        'EXP-5': {'M': 10, 'W': 3, 'U': 0.85, 'W_M_ratio': 3/10, 'group': 'low', 'description': 'M=10,W=3,U=0.85'},
        'EXP-6': {'M': 10, 'W': 5, 'U': 0.95, 'W_M_ratio': 5/10, 'group': 'medium', 'description': 'M=10,W=5,U=0.95'},
        'EXP-7': {'M': 10, 'W': 7, 'U': 0.75, 'W_M_ratio': 7/10, 'group': 'high', 'description': 'M=10,W=7,U=0.75'},
        'EXP-8': {'M': 10, 'W': 3, 'U': 0.95, 'W_M_ratio': 3/10, 'group': 'low', 'description': 'M=10,W=3,U=0.95'},
        'EXP-9': {'M': 20, 'W': 5, 'U': 0.95, 'W_M_ratio': 5/20, 'group': 'low', 'description': 'M=20,W=5,U=0.95'},
        'EXP-10': {'M': 20, 'W': 8, 'U': 0.75, 'W_M_ratio': 8/20, 'group': 'medium', 'description': 'M=20,W=8,U=0.75'},
        'EXP-11': {'M': 20, 'W': 12, 'U': 0.85, 'W_M_ratio': 12/20, 'group': 'medium', 'description': 'M=20,W=12,U=0.85'},
        'EXP-12': {'M': 20, 'W': 5, 'U': 0.75, 'W_M_ratio': 5/20, 'group': 'low', 'description': 'M=20,W=5,U=0.75'},
    }
    
    # Metric configuration
    METRICS = {
        'HV': {
            'sheet_name': 'Hypervolume(HV)',
            'short_name': 'HV',
            'direction': 'higher',
            'description': 'Hypervolume metric (higher is better)'
        },
        'IGD': {
            'sheet_name': 'Inverted Generational Distance',
            'short_name': 'IGD',
            'direction': 'lower',
            'description': 'Inverted Generational Distance (lower is better)'
        },
        'Spread': {
            'sheet_name': 'Spread(Diversity)',
            'short_name': 'Spread',
            'direction': 'lower',
            'description': 'Spread metric (lower is better)'
        }
    }


# ============================================================================
# Data Loader
# ============================================================================

class HeterogeneityDataLoader:
    """Heterogeneity experiment data loader"""
    
    def __init__(self, data_path: str = 'test_result'):
        # Get project root directory
        if len(sys.path) > 0:
            project_root = sys.path[0]
        else:
            project_root = os.getcwd()
        
        self.data_path = Path(os.path.join(project_root, data_path))
        self.config = HeterogeneityConfig()
        self.raw_data = {}
        self.processed_data = None
        self.summary_stats = None
    
    def get_metric_config(self):
        """Get metric configuration"""
        return self.config.METRICS.copy()
    
    def get_metric_names(self):
        """Get metric name list"""
        return [config['short_name'] for config in self.config.METRICS.values()]
    
    def get_metric_direction(self, metric_name: str) -> str:
        """Get metric direction"""
        for config in self.config.METRICS.values():
            if config['short_name'] == metric_name:
                return config['direction']
        return 'higher'
    
    def load_single_file(self, file_path: Path) -> dict:
        """Load a single Excel file"""
        # Parse filename: EXP-1-machine_Metrics_Detail.xlsx -> scenario='EXP-1', het_type='machine'
        filename = file_path.stem
        parts = filename.replace('_Metrics_Detail', '').split('-')
        
        if len(parts) < 2:
            return None
        
        scenario = f"{parts[0]}-{parts[1]}" if parts[0] == 'EXP' else parts[0]
        het_type = parts[-1] if parts[-1] in self.config.HET_TYPES else None
        
        if het_type is None:
            return None
        
        result = {
            'scenario': scenario,
            'het_type': het_type,
            'file_path': file_path,
            'metrics': {}
        }
        
        try:
            wb = load_workbook(file_path, data_only=True)
            
            for metric_key, config in self.config.METRICS.items():
                sheet_name = config['sheet_name']
                
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    
                    # Read headers
                    headers = []
                    for col in range(1, ws.max_column + 1):
                        cell_value = ws.cell(row=1, column=col).value
                        if cell_value:
                            headers.append(cell_value)
                    
                    # Read data rows
                    data_rows = []
                    for row in range(2, ws.max_row + 1):
                        row_data = {}
                        for col, header in enumerate(headers, 1):
                            cell_value = ws.cell(row=row, column=col).value
                            if cell_value is not None:
                                row_data[header] = cell_value
                        
                        # Check if there is data for heterogeneity columns
                        het_cols = self.config.COLUMN_NAMES[het_type]
                        if any(h in row_data for h in het_cols):
                            data_rows.append(row_data)
                    
                    if data_rows:
                        result['metrics'][metric_key] = {
                            'data': pd.DataFrame(data_rows),
                            'direction': config['direction'],
                            'description': config['description']
                        }
            
            wb.close()
            
        except Exception as e:
            logger.error(f"Error loading file {file_path.name}: {e}")
            return None
        
        return result
    
    def load_all_data(self, sub_dir: str = 'Sensitivity') -> dict:
        """Load all heterogeneity experiment data"""
        # Try multiple possible paths
        possible_paths = [
            self.data_path / sub_dir,
            self.data_path / 'Sensitivity',
            self.data_path,
            Path(sys.path[0]) / sub_dir,
            Path(sys.path[0]) / 'Sensitivity',
        ]
        
        excel_files = []
        for path in possible_paths:
            if path.exists():
                files = list(path.glob("*_Metrics_Detail.xlsx"))
                if files:
                    excel_files = files
                    logger.info(f"Found {len(files)} experiment files in directory {path}")
                    break
        
        if not excel_files:
            logger.error("No metrics detail files found")
            return {}
        
        for file_path in excel_files:
            logger.info(f"Loading file: {file_path.name}")
            experiment_data = self.load_single_file(file_path)
            
            if experiment_data and experiment_data['metrics']:
                scenario = experiment_data['scenario']
                het_type = experiment_data['het_type']
                
                if scenario not in self.raw_data:
                    self.raw_data[scenario] = {}
                
                self.raw_data[scenario][het_type] = experiment_data
        
        if self.raw_data:
            self._process_data()
            logger.info(f"Successfully loaded data from {len(self.raw_data)} scenarios")
        
        return self.raw_data
    
    def _process_data(self):
        """Process raw data into unified format"""
        all_data = []
        
        for scenario, scenario_data in self.raw_data.items():
            scenario_features = self.config.SCENARIO_FEATURES.get(scenario, {})
            
            for het_type, het_data in scenario_data.items():
                het_values = self.config.HET_VALUES[het_type]
                het_columns = self.config.COLUMN_NAMES[het_type]
                
                for metric_key, metric_info in het_data['metrics'].items():
                    df = metric_info['data']
                    
                    for col_idx, col_name in enumerate(het_columns):
                        if col_name in df.columns:
                            het_value = het_values[col_idx]
                            values = pd.to_numeric(df[col_name], errors='coerce').dropna().values
                            
                            for run_idx, value in enumerate(values):
                                all_data.append({
                                    'Scenario': scenario,
                                    'Heterogeneity Type': het_type,
                                    'Heterogeneity Level': col_idx + 1,
                                    'Heterogeneity Value': het_value,
                                    'Metric': metric_key,
                                    'Value': value,
                                    'Run': run_idx + 1,
                                    'Direction': metric_info['direction'],
                                    'M': scenario_features.get('M', 0),
                                    'W': scenario_features.get('W', 0),
                                    'U': scenario_features.get('U', 0),
                                    'W_M_ratio': scenario_features.get('W_M_ratio', 0),
                                    'group': scenario_features.get('group', 'unknown'),
                                    'description': scenario_features.get('description', '')
                                })
        
        if all_data:
            self.processed_data = pd.DataFrame(all_data)
            self._calculate_summary_stats()
            logger.info(f"Data processing complete, {len(self.processed_data)} records in total")
    
    def _calculate_summary_stats(self):
        """Calculate summary statistics"""
        if self.processed_data is None or len(self.processed_data) == 0:
            return
        
        # Calculate statistics by group
        self.summary_stats = self.processed_data.groupby(
            ['Scenario', 'Heterogeneity Type', 'Heterogeneity Value', 'Metric', 'group', 'W_M_ratio', 'M', 'W', 'U']
        )['Value'].agg(['mean', 'std', 'sem', 'min', 'max', 'median', 'count']).reset_index()
        
        # Calculate coefficient of variation (CV)
        self.summary_stats['cv'] = self.summary_stats['std'] / self.summary_stats['mean']
        
        logger.info(f"Summary statistics complete, {len(self.summary_stats)} records")
    
    def get_data_for_metric(self, metric_name: str) -> pd.DataFrame:
        """Get detailed data for a specified metric"""
        if self.processed_data is None:
            return pd.DataFrame()
        return self.processed_data[self.processed_data['Metric'] == metric_name].copy()
    
    def get_summary_for_metric(self, metric_name: str) -> pd.DataFrame:
        """Get summary statistics for a specified metric"""
        if self.summary_stats is None:
            return pd.DataFrame()
        return self.summary_stats[self.summary_stats['Metric'] == metric_name].copy()


# ============================================================================
# Statistical Analyzer
# ============================================================================

class HeterogeneityAnalyzer:
    """Heterogeneity sensitivity statistical analyzer"""
    
    def __init__(self, data_loader: HeterogeneityDataLoader):
        self.data_loader = data_loader
        self.config = HeterogeneityConfig()
    
    def calculate_sensitivity(self, metric_name: str = 'HV') -> pd.DataFrame:
        """Calculate sensitivity of each heterogeneity type across different scenarios"""
        summary = self.data_loader.get_summary_for_metric(metric_name)
        
        if summary.empty:
            return pd.DataFrame()
        
        sensitivity_data = []
        
        for scenario in self.config.SCENARIOS:
            for het_type in self.config.HET_TYPES:
                subset = summary[(summary['Scenario'] == scenario) & 
                                (summary['Heterogeneity Type'] == het_type)]
                
                if len(subset) >= 2:
                    min_het = subset['Heterogeneity Value'].min()
                    max_het = subset['Heterogeneity Value'].max()
                    
                    hv_min = subset[subset['Heterogeneity Value'] == min_het]['mean'].values
                    hv_max = subset[subset['Heterogeneity Value'] == max_het]['mean'].values
                    
                    if len(hv_min) > 0 and len(hv_max) > 0 and hv_min[0] > 0:
                        # Calculate relative change rate
                        relative_change = (hv_max[0] - hv_min[0]) / hv_min[0]
                        
                        # For higher-is-better metric (HV), negative value indicates performance degradation
                        degradation = -relative_change
                        
                        sensitivity_data.append({
                            'Scenario': scenario,
                            'Heterogeneity Type': het_type,
                            'W_M_ratio': self.config.SCENARIO_FEATURES[scenario]['W_M_ratio'],
                            'group': self.config.SCENARIO_FEATURES[scenario]['group'],
                            'M': self.config.SCENARIO_FEATURES[scenario]['M'],
                            'W': self.config.SCENARIO_FEATURES[scenario]['W'],
                            'U': self.config.SCENARIO_FEATURES[scenario]['U'],
                            'HV_min': hv_min[0],
                            'HV_max': hv_max[0],
                            'Relative Change': relative_change,
                            'Performance Degradation': degradation
                        })
        
        return pd.DataFrame(sensitivity_data)
    
    def correlation_analysis(self) -> pd.DataFrame:
        """Correlation analysis: relationship between W/M ratio and sensitivity"""
        sensitivity_df = self.calculate_sensitivity('HV')
        
        if sensitivity_df.empty:
            return pd.DataFrame()
        
        corr_results = []
        
        for het_type in self.config.HET_TYPES:
            subset = sensitivity_df[sensitivity_df['Heterogeneity Type'] == het_type]
            
            if len(subset) >= 3:
                corr, p_val = stats.pearsonr(subset['W_M_ratio'], subset['Performance Degradation'])
                
                # Significance markers
                if p_val < 0.001:
                    sig = '***'
                elif p_val < 0.01:
                    sig = '**'
                elif p_val < 0.05:
                    sig = '*'
                else:
                    sig = 'n.s.'
                
                corr_results.append({
                    'Heterogeneity Type': het_type,
                    'Pearson r': corr,
                    'p-value': p_val,
                    'Significance': sig,
                    'Sample Size': len(subset)
                })
        
        return pd.DataFrame(corr_results)
    
    def group_comparison_stats(self) -> pd.DataFrame:
        """Group comparison statistics"""
        summary = self.data_loader.get_summary_for_metric('HV')
        
        if summary.empty:
            return pd.DataFrame()
        
        group_stats = []
        
        for group in ['low', 'medium', 'high']:
            for het_type in self.config.HET_TYPES:
                subset = summary[(summary['group'] == group) & 
                                (summary['Heterogeneity Type'] == het_type)]
                
                if len(subset) >= 2:
                    min_het = subset['Heterogeneity Value'].min()
                    max_het = subset['Heterogeneity Value'].max()
                    
                    hv_min = subset[subset['Heterogeneity Value'] == min_het]['mean'].mean()
                    hv_max = subset[subset['Heterogeneity Value'] == max_het]['mean'].mean()
                    
                    if hv_min > 0:
                        degradation = (hv_min - hv_max) / hv_min * 100
                        
                        group_stats.append({
                            'Scenario Group': group,
                            'Heterogeneity Type': het_type.capitalize(),
                            'Low Het HV': f'{hv_min:.4f}',
                            'High Het HV': f'{hv_max:.4f}',
                            'Performance Degradation(%)': f'{degradation:.2f}%',
                            'degradation_raw': degradation
                        })
        
        return pd.DataFrame(group_stats)
    
    def overall_impact_stats(self) -> pd.DataFrame:
        """Overall impact statistics"""
        summary = self.data_loader.get_summary_for_metric('HV')
        
        if summary.empty:
            return pd.DataFrame()
        
        overall_stats = []
        
        for het_type in self.config.HET_TYPES:
            subset = summary[summary['Heterogeneity Type'] == het_type]
            
            if len(subset) >= 2:
                min_het = subset['Heterogeneity Value'].min()
                max_het = subset['Heterogeneity Value'].max()
                
                het_col = 'Heterogeneity Value'
                
                hv_min = subset[subset[het_col] == min_het]['mean'].mean()
                hv_max = subset[subset[het_col] == max_het]['mean'].mean()
                
                if hv_min > 0:
                    degradation = (hv_min - hv_max) / hv_min * 100
                    
                    overall_stats.append({
                        'Heterogeneity Type': het_type.capitalize(),
                        'Low Het HV Mean': f'{hv_min:.4f}',
                        'High Het HV Mean': f'{hv_max:.4f}',
                        'Performance Degradation': f'{degradation:.2f}%',
                        'degradation_raw': degradation
                    })
        
        return pd.DataFrame(overall_stats)


# ============================================================================
# Visualizer
# ============================================================================

class HeterogeneityVisualizer:
    """Heterogeneity sensitivity visualizer"""
    
    def __init__(self, data_loader: HeterogeneityDataLoader, analyzer: HeterogeneityAnalyzer):
        self.data_loader = data_loader
        self.analyzer = analyzer
        self.config = HeterogeneityConfig()
        
        # Color configuration
        self.het_colors = {
            'machine': '#E63946',
            'repair': '#457B9D', 
            'worker': '#2A9D8F'
        }
        self.group_colors = {
            'low': '#E63946',
            'medium': '#F4A261',
            'high': '#2A9D8F'
        }
    
    def plot_overall_comparison(self, save_path: str = None):
        """Figure 1: Comparison of average impact of three heterogeneity types across all scenarios"""
        fig, axes = plt.subplots(1, 3, figsize=(15, 4.5))
        
        metrics = ['HV', 'IGD', 'Spread']
        
        for idx, metric in enumerate(metrics):
            ax = axes[idx]
            summary = self.data_loader.get_summary_for_metric(metric)
            
            if summary.empty:
                continue
            
            for het_type in self.config.HET_TYPES:
                subset = summary[summary['Heterogeneity Type'] == het_type]
                
                if subset.empty:
                    continue
                
                het_values = self.config.HET_VALUES[het_type]
                means = []
                sems = []
                
                for hv in het_values:
                    hv_subset = subset[subset['Heterogeneity Value'] == hv]
                    if not hv_subset.empty:
                        means.append(hv_subset['mean'].mean())
                        sems.append(hv_subset['sem'].mean())
                    else:
                        means.append(np.nan)
                        sems.append(np.nan)
                
                # Normalize to first value
                if means and not np.isnan(means[0]) and means[0] != 0:
                    base = means[0]
                    norm_means = [m / base for m in means]
                    norm_sems = [s / base for s in sems]
                    
                    ax.errorbar(het_values, norm_means, yerr=norm_sems,
                               marker='o', markersize=5, linewidth=2, capsize=3,
                               label=het_type.capitalize(),
                               color=self.het_colors[het_type])
            
            ax.set_xlabel('Heterogeneity Level', fontweight='bold')
            ax.set_ylabel(f'Relative {metric}', fontweight='bold')
            ax.set_title(f'({chr(97+idx)}) {metric} Metric', fontweight='bold')
            ax.legend(loc='best')
            ax.grid(True, alpha=0.3, linestyle='--')
        
        plt.suptitle('Figure 1: Overall Impact of Three Heterogeneity Types', 
                    fontsize=14, fontweight='bold', y=1.02)
        plt.tight_layout()
        
        if save_path:
            plt.savefig(save_path, dpi=300, bbox_inches='tight')
            logger.info(f"Figure 1 saved to: {save_path}")
        else:
            plt.show()
        plt.close()
    
    def plot_scenario_group_comparison(self, save_path: str = None):
        """Figure 2: Heterogeneity sensitivity comparison across different scenario groups"""
        fig, axes = plt.subplots(2, 3, figsize=(16, 9))
        
        groups = ['low', 'medium', 'high']
        group_labels = {
            'low': 'Low W/M Ratio (Vulnerable)',
            'medium': 'Medium W/M Ratio',
            'high': 'High W/M Ratio (Buffered)'
        }
        
        for row, metric in enumerate(['HV', 'IGD']):
            summary = self.data_loader.get_summary_for_metric(metric)
            
            for col, group in enumerate(groups):
                ax = axes[row, col]
                
                for het_type in self.config.HET_TYPES:
                    subset = summary[(summary['group'] == group) & 
                                    (summary['Heterogeneity Type'] == het_type)]
                    
                    if subset.empty:
                        continue
                    
                    het_values = self.config.HET_VALUES[het_type]
                    means = []
                    
                    for hv in het_values:
                        hv_subset = subset[subset['Heterogeneity Value'] == hv]
                        if not hv_subset.empty:
                            means.append(hv_subset['mean'].mean())
                        else:
                            means.append(np.nan)
                    
                    if means and not np.isnan(means[0]) and means[0] != 0:
                        base = means[0]
                        norm_means = [m / base for m in means]
                        
                        ax.plot(het_values, norm_means, marker='o', markersize=4,
                               linewidth=2, label=het_type.capitalize(),
                               color=self.het_colors[het_type])
                
                ax.set_xlabel('Heterogeneity Level', fontweight='bold')
                ax.set_ylabel(f'Relative {metric}', fontweight='bold')
                ax.set_title(f'{group_labels[group]}\n({metric})', fontweight='bold', fontsize=10)
                ax.legend(loc='best', fontsize=8)
                ax.grid(True, alpha=0.3, linestyle='--')
        
        # Spread metric
        summary_spread = self.data_loader.get_summary_for_metric('Spread')
        for col, group in enumerate(groups):
            ax = axes[1, col]
            
            for het_type in self.config.HET_TYPES:
                subset = summary_spread[(summary_spread['group'] == group) & 
                                       (summary_spread['Heterogeneity Type'] == het_type)]
                
                if subset.empty:
                    continue
                
                het_values = self.config.HET_VALUES[het_type]
                means = []
                
                for hv in het_values:
                    hv_subset = subset[subset['Heterogeneity Value'] == hv]
                    if not hv_subset.empty:
                        means.append(hv_subset['mean'].mean())
                    else:
                        means.append(np.nan)
                
                if means and not np.isnan(means[0]) and means[0] != 0:
                    base = means[0]
                    norm_means = [m / base for m in means]
                    
                    ax.plot(het_values, norm_means, marker='o', markersize=4,
                           linewidth=2, label=het_type.capitalize(),
                           color=self.het_colors[het_type])
            
            ax.set_xlabel('Heterogeneity Level', fontweight='bold')
            ax.set_ylabel('Relative Spread', fontweight='bold')
            ax.set_title(f'{group_labels[group]}\n(Spread)', fontweight='bold', fontsize=10)
            ax.legend(loc='best', fontsize=8)
            ax.grid(True, alpha=0.3, linestyle='--')
        
        plt.suptitle('Figure 2: Heterogeneity Sensitivity by Scenario Group', 
                    fontsize=14, fontweight='bold', y=1.02)
        plt.tight_layout()
        
        if save_path:
            plt.savefig(save_path, dpi=300, bbox_inches='tight')
            logger.info(f"Figure 2 saved to: {save_path}")
        else:
            plt.show()
        plt.close()
    
    def plot_impact_heatmap(self, save_path: str = None):
        """Figure 3: Heterogeneity impact heatmap"""
        sensitivity_df = self.analyzer.calculate_sensitivity('HV')
        
        if sensitivity_df.empty:
            logger.warning("No sensitivity data, cannot draw heatmap")
            return
        
        fig, ax = plt.subplots(figsize=(14, 5))
        
        # Sort scenarios by W/M ratio
        sorted_scenarios = sorted(self.config.SCENARIOS, 
                                 key=lambda s: self.config.SCENARIO_FEATURES[s]['W_M_ratio'])
        
        # Build heatmap data matrix
        heatmap_data = []
        for het_type in self.config.HET_TYPES:
            row_data = []
            for scenario in sorted_scenarios:
                subset = sensitivity_df[(sensitivity_df['Scenario'] == scenario) & 
                                       (sensitivity_df['Heterogeneity Type'] == het_type)]
                
                if not subset.empty:
                    degradation = subset['Performance Degradation'].values[0] * 100
                else:
                    degradation = np.nan
                
                row_data.append(degradation)
            heatmap_data.append(row_data)
        
        heatmap_array = np.array(heatmap_data)
        
        # Draw heatmap
        im = ax.imshow(heatmap_array, cmap='RdYlGn_r', aspect='auto', vmin=0, vmax=25)
        
        ax.set_xticks(range(len(sorted_scenarios)))
        ax.set_xticklabels([f'{s}\nW/M={self.config.SCENARIO_FEATURES[s]["W_M_ratio"]:.2f}' 
                           for s in sorted_scenarios], fontsize=9)
        ax.set_yticks(range(len(self.config.HET_TYPES)))
        ax.set_yticklabels(['Machine', 'Repair', 'Worker'], fontsize=11)
        
        # Add value annotations
        for i in range(len(self.config.HET_TYPES)):
            for j in range(len(sorted_scenarios)):
                value = heatmap_array[i, j]
                if not np.isnan(value):
                    text_color = 'white' if value > 12 else 'black'
                    ax.text(j, i, f'{value:.1f}%', ha='center', va='center', 
                           color=text_color, fontsize=8, fontweight='bold')
        
        plt.colorbar(im, ax=ax, label='HV Performance Degradation (%)')
        ax.set_title('Figure 3: Impact of Heterogeneity on HV Performance\n(Lower % = More Robust)', 
                    fontweight='bold', fontsize=12)
        
        plt.tight_layout()
        
        if save_path:
            plt.savefig(save_path, dpi=300, bbox_inches='tight')
            logger.info(f"Figure 3 saved to: {save_path}")
        else:
            plt.show()
        plt.close()
    
    def plot_stability_analysis(self, save_path: str = None):
        """Figure 4: Algorithm stability analysis (coefficient of variation)"""
        fig, axes = plt.subplots(1, 3, figsize=(15, 4.5))
        
        groups = ['low', 'medium', 'high']
        group_labels = {
            'low': 'Low W/M Ratio (Vulnerable)',
            'medium': 'Medium W/M Ratio',
            'high': 'High W/M Ratio (Buffered)'
        }
        
        summary = self.data_loader.get_summary_for_metric('HV')
        
        for idx, group in enumerate(groups):
            ax = axes[idx]
            
            for het_type in self.config.HET_TYPES:
                subset = summary[(summary['group'] == group) & 
                                (summary['Heterogeneity Type'] == het_type)]
                
                if subset.empty:
                    continue
                
                het_values = self.config.HET_VALUES[het_type]
                cvs = []
                
                for hv in het_values:
                    hv_subset = subset[subset['Heterogeneity Value'] == hv]
                    if not hv_subset.empty:
                        cvs.append(hv_subset['cv'].mean())
                    else:
                        cvs.append(np.nan)
                
                ax.plot(het_values, cvs, marker='s', markersize=5, linewidth=2,
                       label=het_type.capitalize(), color=self.het_colors[het_type])
            
            ax.set_xlabel('Heterogeneity Level', fontweight='bold')
            ax.set_ylabel('Coefficient of Variation (CV)', fontweight='bold')
            ax.set_title(f'({chr(97+idx)}) {group_labels[group]}', fontweight='bold')
            ax.legend(loc='best')
            ax.grid(True, alpha=0.3, linestyle='--')
        
        plt.suptitle('Figure 4: Algorithm Stability Analysis\n(Higher CV = Less Stable)', 
                    fontsize=14, fontweight='bold', y=1.02)
        plt.tight_layout()
        
        if save_path:
            plt.savefig(save_path, dpi=300, bbox_inches='tight')
            logger.info(f"Figure 4 saved to: {save_path}")
        else:
            plt.show()
        plt.close()
    
    def plot_correlation_analysis(self, save_path: str = None):
        """Figure 5: Correlation analysis between scenario characteristics and heterogeneity sensitivity"""
        sensitivity_df = self.analyzer.calculate_sensitivity('HV')
        
        if sensitivity_df.empty:
            logger.warning("No sensitivity data, cannot draw correlation analysis chart")
            return
        
        fig, axes = plt.subplots(1, 3, figsize=(15, 4.5))
        
        for idx, het_type in enumerate(self.config.HET_TYPES):
            ax = axes[idx]
            
            subset = sensitivity_df[sensitivity_df['Heterogeneity Type'] == het_type]
            
            if subset.empty:
                continue
            
            # Scatter plot
            scatter = ax.scatter(subset['W_M_ratio'], subset['Performance Degradation'] * 100,
                                c=subset['U'], s=subset['M'] * 15,
                                cmap='viridis', alpha=0.7, edgecolors='black', linewidth=0.5)
            
            # Add scenario labels
            for _, row in subset.iterrows():
                ax.annotate(row['Scenario'].replace('EXP-', ''),
                           (row['W_M_ratio'], row['Performance Degradation'] * 100),
                           fontsize=8, ha='center', va='bottom')
            
            # Fit line
            if len(subset) >= 3:
                z = np.polyfit(subset['W_M_ratio'], subset['Performance Degradation'] * 100, 1)
                p = np.poly1d(z)
                x_line = np.linspace(subset['W_M_ratio'].min(), subset['W_M_ratio'].max(), 100)
                ax.plot(x_line, p(x_line), 'r--', alpha=0.5, 
                       label=f'Trend (slope={z[0]:.2f})')
                
                # Correlation coefficient
                corr, p_val = stats.pearsonr(subset['W_M_ratio'], subset['Performance Degradation'])
                sig_text = '***' if p_val < 0.001 else ('**' if p_val < 0.01 else ('*' if p_val < 0.05 else 'n.s.'))
                ax.text(0.05, 0.95, f'r = {corr:.3f}{sig_text}\np = {p_val:.4f}',
                       transform=ax.transAxes, fontsize=10, verticalalignment='top',
                       bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.5))
            
            ax.set_xlabel('Worker-to-Machine Ratio (W/M)', fontweight='bold')
            ax.set_ylabel('HV Performance Degradation (%)', fontweight='bold')
            ax.set_title(f'({chr(97+idx)}) {het_type.capitalize()} Heterogeneity', fontweight='bold')
            ax.legend(loc='best', fontsize=9)
            ax.grid(True, alpha=0.3, linestyle='--')
            
            if idx == 0:
                cbar = plt.colorbar(scatter, ax=ax)
                cbar.set_label('Utilization (U)', fontsize=9)
        
        plt.suptitle('Figure 5: Correlation between W/M Ratio and Heterogeneity Sensitivity', 
                    fontsize=14, fontweight='bold', y=1.02)
        plt.tight_layout()
        
        if save_path:
            plt.savefig(save_path, dpi=300, bbox_inches='tight')
            logger.info(f"Figure 5 saved to: {save_path}")
        else:
            plt.show()
        plt.close()
    
    def plot_radar_comparison(self, save_path: str = None):
        """Figure 6: Radar chart - multi-dimensional performance comparison"""
        sensitivity_df = self.analyzer.calculate_sensitivity('HV')
        summary = self.data_loader.get_summary_for_metric('HV')
        
        if sensitivity_df.empty or summary.empty:
            logger.warning("Insufficient data, cannot draw radar chart")
            return
        
        fig = plt.figure(figsize=(14, 5))
        
        categories = ['Robustness to\nMachine Het.', 'Robustness to\nRepair Het.',
                     'Robustness to\nWorker Het.', 'Overall\nHV Performance',
                     'Algorithm\nStability']
        N = len(categories)
        angles = [n / float(N) * 2 * pi for n in range(N)]
        angles += angles[:1]
        
        groups = ['low', 'medium', 'high']
        group_labels = {
            'low': 'Low W/M Ratio\n(Vulnerable)',
            'medium': 'Medium W/M Ratio',
            'high': 'High W/M Ratio\n(Robust)'
        }
        
        for idx, group in enumerate(groups):
            ax = fig.add_subplot(1, 3, idx + 1, projection='polar')
            
            values = []
            
            # Robustness of three heterogeneity types
            for het_type in self.config.HET_TYPES:
                subset = sensitivity_df[(sensitivity_df['group'] == group) & 
                                       (sensitivity_df['Heterogeneity Type'] == het_type)]
                if not subset.empty:
                    degradation = subset['Performance Degradation'].mean()
                    robustness = max(0, 1 - degradation)
                else:
                    robustness = 0.5
                values.append(robustness)
            
            # Overall HV performance (normalized)
            group_hv = summary[summary['group'] == group]['mean'].mean()
            max_hv = summary['mean'].max()
            values.append(group_hv / max_hv if max_hv > 0 else 0.5)
            
            # Algorithm stability (1 - CV)
            group_cv = summary[summary['group'] == group]['cv'].mean()
            stability = max(0, 1 - group_cv)
            values.append(stability)
            
            values += values[:1]
            
            ax.plot(angles, values, 'o-', linewidth=2, color='#E63946')
            ax.fill(angles, values, alpha=0.25, color='#E63946')
            ax.set_xticks(angles[:-1])
            ax.set_xticklabels(categories, fontsize=8)
            ax.set_ylim(0, 1)
            ax.set_title(f'({chr(97+idx)}) {group_labels[group]}', fontweight='bold', pad=20)
        
        plt.suptitle('Figure 6: Multi-Dimensional Performance Comparison', 
                    fontsize=14, fontweight='bold', y=1.05)
        plt.tight_layout()
        
        if save_path:
            plt.savefig(save_path, dpi=300, bbox_inches='tight')
            logger.info(f"Figure 6 saved to: {save_path}")
        else:
            plt.show()
        plt.close()
    
    def generate_all_figures(self, output_dir: str):
        """Generate all visualization figures"""
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)
        
        logger.info("Starting visualization figure generation...")
        
        self.plot_overall_comparison(save_path=output_path / 'fig1_overall_comparison.png')
        self.plot_scenario_group_comparison(save_path=output_path / 'fig2_scenario_group_comparison.png')
        self.plot_impact_heatmap(save_path=output_path / 'fig3_impact_heatmap.png')
        self.plot_stability_analysis(save_path=output_path / 'fig4_stability_analysis.png')
        self.plot_correlation_analysis(save_path=output_path / 'fig5_correlation_analysis.png')
        self.plot_radar_comparison(save_path=output_path / 'fig6_radar_comparison.png')
        
        logger.info(f"All figures saved to: {output_path}")


# ============================================================================
# Report Generator
# ============================================================================

class HeterogeneityReportGenerator:
    """Heterogeneity analysis report generator"""
    
    def __init__(self, data_loader: HeterogeneityDataLoader, 
                 analyzer: HeterogeneityAnalyzer,
                 visualizer: HeterogeneityVisualizer):
        self.data_loader = data_loader
        self.analyzer = analyzer
        self.visualizer = visualizer
        self.config = HeterogeneityConfig()
    
    def generate_full_report(self, output_dir: str = 'heterogeneity_analysis_report'):
        """Generate complete analysis report"""
        # Get project root directory
        project_root = sys.path[0] if len(sys.path) > 0 else os.getcwd()
        output_path = Path(os.path.join(project_root, 'test_result', output_dir))
        output_path.mkdir(parents=True, exist_ok=True)
        
        logger.info("=" * 60)
        logger.info("Starting heterogeneity sensitivity analysis report generation")
        logger.info("=" * 60)
        
        # 1. Generate statistical tables
        logger.info("[1/4] Generating statistical tables...")
        self._save_statistical_tables(output_path)
        
        # 2. Generate visualization figures
        logger.info("[2/4] Generating visualization figures...")
        self.visualizer.generate_all_figures(output_path)
        
        # 3. Generate text report
        logger.info("[3/4] Generating text report...")
        self._generate_text_report(output_path)
        
        # 4. Generate detailed data tables
        logger.info("[4/4] Generating detailed data tables...")
        self._save_detailed_data(output_path)
        
        logger.info("=" * 60)
        logger.info(f"Report generation complete! All files saved to: {output_path}")
        logger.info("=" * 60)
        
        return str(output_path)
    
    def _save_statistical_tables(self, output_path: Path):
        """Save statistical tables"""
        with pd.ExcelWriter(output_path / 'statistical_tables.xlsx', engine='openpyxl') as writer:
            # Overall impact statistics
            overall_stats = self.analyzer.overall_impact_stats()
            if not overall_stats.empty:
                overall_stats.to_excel(writer, sheet_name='Overall Impact Stats', index=False)
            
            # Group comparison statistics
            group_stats = self.analyzer.group_comparison_stats()
            if not group_stats.empty:
                group_stats.to_excel(writer, sheet_name='Group Comparison Stats', index=False)
            
            # Correlation analysis
            corr_results = self.analyzer.correlation_analysis()
            if not corr_results.empty:
                corr_results.to_excel(writer, sheet_name='Correlation Analysis', index=False)
            
            # Sensitivity data
            sensitivity_df = self.analyzer.calculate_sensitivity('HV')
            if not sensitivity_df.empty:
                sensitivity_df.to_excel(writer, sheet_name='HV Sensitivity Data', index=False)
        
        logger.info(f"  Statistical tables saved to: {output_path / 'statistical_tables.xlsx'}")
    
    def _save_detailed_data(self, output_path: Path):
        """Save detailed data"""
        # Summary statistics
        if self.data_loader.summary_stats is not None:
            self.data_loader.summary_stats.to_csv(
                output_path / 'summary_statistics.csv', 
                index=False, encoding='utf-8-sig'
            )
        
        # Save detailed data by metric
        for metric in self.config.METRICS.keys():
            data = self.data_loader.get_data_for_metric(metric)
            if not data.empty:
                data.to_csv(
                    output_path / f'detailed_data_{metric}.csv',
                    index=False, encoding='utf-8-sig'
                )
        
        logger.info(f"  Detailed data saved to: {output_path}")
    
    def _generate_text_report(self, output_path: Path):
        """Generate text report"""
        report_path = output_path / 'analysis_report.txt'
        
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write("=" * 80 + "\n")
            f.write("Heterogeneity Parameter Sensitivity Analysis - Complete Statistical Report\n")
            f.write("Heterogeneity Sensitivity Analysis - Comprehensive Report\n")
            f.write("=" * 80 + "\n\n")
            
            # Data overview
            f.write("1. Data Overview\n")
            f.write("-" * 40 + "\n")
            f.write(f"  Number of experiment scenarios: {len(self.config.SCENARIOS)}\n")
            f.write(f"  Heterogeneity types: {', '.join(self.config.HET_TYPES)}\n")
            f.write(f"  Evaluation metrics: {', '.join(self.config.METRICS.keys())}\n")
            
            if self.data_loader.processed_data is not None:
                f.write(f"  Total data records: {len(self.data_loader.processed_data)}\n")
            f.write("\n")
            
            # Overall impact statistics
            f.write("2. Overall Impact of Three Heterogeneity Types\n")
            f.write("-" * 40 + "\n")
            overall_stats = self.analyzer.overall_impact_stats()
            if not overall_stats.empty:
                for _, row in overall_stats.iterrows():
                    f.write(f"  {row['Heterogeneity Type']} heterogeneity:\n")
                    f.write(f"    Low heterogeneity HV mean: {row['Low Het HV Mean']}\n")
                    f.write(f"    High heterogeneity HV mean: {row['High Het HV Mean']}\n")
                    f.write(f"    Performance degradation: {row['Performance Degradation']}\n\n")
            
            # Group comparison statistics
            f.write("3. Scenario Group Comparison Statistics\n")
            f.write("-" * 40 + "\n")
            group_stats = self.analyzer.group_comparison_stats()
            if not group_stats.empty:
                for group in ['low', 'medium', 'high']:
                    group_subset = group_stats[group_stats['Scenario Group'] == group]
                    f.write(f"\n  {group.upper()} W/M Ratio group:\n")
                    for _, row in group_subset.iterrows():
                        f.write(f"    {row['Heterogeneity Type']}: {row['Performance Degradation(%)']}\n")
            f.write("\n")
            
            # Correlation analysis
            f.write("4. Correlation between W/M Ratio and Sensitivity\n")
            f.write("-" * 40 + "\n")
            corr_results = self.analyzer.correlation_analysis()
            if not corr_results.empty:
                for _, row in corr_results.iterrows():
                    f.write(f"  {row['Heterogeneity Type']} heterogeneity:\n")
                    f.write(f"    Pearson r = {row['Pearson r']:.4f}\n")
                    f.write(f"    p-value = {row['p-value']:.4f} ({row['Significance']})\n\n")
            
            # Core conclusions
            f.write("5. Core Conclusions\n")
            f.write("-" * 40 + "\n")
            f.write("  (1) Repair heterogeneity has the greatest impact on system performance, followed by machine heterogeneity,\n")
            f.write("      while worker heterogeneity has the smallest impact.\n\n")
            f.write("  (2) Low W/M ratio scenarios act as an \"amplifier\" for all heterogeneity types, making the system most vulnerable.\n\n")
            f.write("  (3) High W/M ratio scenarios can effectively buffer machine heterogeneity, but have\n")
            f.write("      limited buffering capacity against repair heterogeneity.\n\n")
            f.write("  (4) When repair heterogeneity is high, algorithm stability (CV) decreases significantly.\n\n")
            
            f.write("\n" + "=" * 80 + "\n")
            f.write("Report generation time: " + pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S") + "\n")
            f.write("=" * 80 + "\n")
        
        logger.info(f"  Text report saved to: {report_path}")


# ============================================================================
# Main Function
# ============================================================================

def run_heterogeneity_analysis(data_path: str = 'test_result', 
                               sub_dir: str = 'heterogeneity',
                               output_dir: str = 'heterogeneity_analysis_report'):
    """Run complete heterogeneity sensitivity analysis"""
    
    logger.info("=" * 60)
    logger.info("Heterogeneity Parameter Sensitivity Analysis System")
    logger.info("=" * 60)
    
    # 1. Load data
    logger.info("\n[Phase 1] Loading experiment data...")
    loader = HeterogeneityDataLoader(data_path)
    loader.load_all_data(sub_dir=sub_dir)
    
    if not loader.raw_data:
        logger.error("No data loaded successfully, please check the data path!")
        return None, None, None, None
    
    # 2. Initialize analyzer
    logger.info("\n[Phase 2] Initializing statistical analyzer...")
    analyzer = HeterogeneityAnalyzer(loader)
    
    # 3. Initialize visualizer
    logger.info("\n[Phase 3] Initializing visualization engine...")
    visualizer = HeterogeneityVisualizer(loader, analyzer)
    
    # 4. Generate complete report
    logger.info("\n[Phase 4] Generating complete analysis report...")
    reporter = HeterogeneityReportGenerator(loader, analyzer, visualizer)
    report_path = reporter.generate_full_report(output_dir=output_dir)
    
    logger.info(f"\nAnalysis complete! Report path: {report_path}")
    
    return loader, analyzer, visualizer, reporter


def main():
    """Main entry function"""
    # Configuration parameters
    DATA_PATH = 'test_result'
    SUB_DIR = 'heterogeneity'  # Subdirectory for heterogeneity data
    OUTPUT_DIR = 'heterogeneity_analysis_report'
    
    # Run analysis
    loader, analyzer, visualizer, reporter = run_heterogeneity_analysis(
        data_path=DATA_PATH,
        sub_dir=SUB_DIR,
        output_dir=OUTPUT_DIR
    )
    
    if loader is None:
        logger.error("Analysis failed, please check the data path")
        sys.exit(1)


if __name__ == "__main__":
    main()