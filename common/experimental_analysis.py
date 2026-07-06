import os
import sys
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook, Workbook
from pathlib import Path
from scipy import stats
from scipy.stats import friedmanchisquare, rankdata, wilcoxon, mannwhitneyu
from scipy.spatial import KDTree
import scikit_posthocs as sp
from typing import Dict, List, Tuple, Optional, Union
import json
from math import pi
import warnings
import logging
warnings.filterwarnings('ignore')

# Configure Chinese display
plt.rcParams['font.sans-serif'] = ['SimHei', 'DejaVu Sans', 'Arial Unicode MS', 'Microsoft YaHei']
plt.rcParams['axes.unicode_minus'] = False
plt.rcParams['figure.dpi'] = 100
plt.rcParams['savefig.dpi'] = 300

# Configure structured logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('experiment_analysis.log', encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)



# ============================================================================
# Common Utility Functions
# ============================================================================
def is_dominated(sol_a: np.ndarray, sol_b: np.ndarray) -> bool:
    """Check if sol_a dominates sol_b (minimization problem)"""
    return np.all(sol_a <= sol_b) and np.any(sol_a < sol_b)


def calculate_nearest_neighbor_distances(solutions: np.ndarray) -> np.ndarray:
    """Calculate distance from each solution to its nearest neighbor (using KDTree)"""
    if len(solutions) <= 1:
        return np.array([])
    
    tree = KDTree(solutions)
    distances, _ = tree.query(solutions, k=2)
    return distances[:, 1]


def validate_solutions(solutions: np.ndarray, name: str = "Solution set") -> None:
    """Validate the validity of a solution set"""
    if not isinstance(solutions, np.ndarray):
        raise TypeError(f"{name} must be a numpy array")
    if solutions.ndim != 2:
        raise ValueError(f"{name} must be a 2D array (num_samples, num_objectives)")
    if np.any(np.isnan(solutions)) or np.any(np.isinf(solutions)):
        raise ValueError(f"{name} contains NaN/Inf values, cannot compute metrics")


# ============================================================================
# Part 1: Multi-Objective Optimization Experiment Data Generation and Management
# ============================================================================
class MultiObjectiveManager:
    """Multi-objective analysis manager - computes HV, IGD, and Spread metrics"""
    
    def __init__(self):
        self.all_experiment_data = {}
        self.global_ideal = None
        self.global_nadir = None
        self.normalization_ranges = None
        self.algorithm_pareto_solutions = {}
        self.global_pareto_front = None
        self.rule_non_dominated_cache = {}
        self.job_pareto_fronts = {}
    
    def add_experiment_data(self, rule_name, run_id, job_objectives):
        """Add experiment data"""
        if rule_name not in self.all_experiment_data:
            self.all_experiment_data[rule_name] = {}
        self.all_experiment_data[rule_name][run_id] = job_objectives
        self.rule_non_dominated_cache = {}
        self.algorithm_pareto_solutions = {}
        self.global_pareto_front = None
        self.job_pareto_fronts = {}
    
    def calculate_global_reference_points(self):
        """Calculate global reference points for normalization"""
        all_solutions = []
        for rule_data in self.all_experiment_data.values():
            for run_data in rule_data.values():
                for job_data in run_data:
                    if 'objectives' in job_data:
                        all_solutions.append(job_data['objectives'])
        
        if all_solutions:
            all_solutions_array = np.array(all_solutions)
            validate_solutions(all_solutions_array, "Global solution set")
            
            self.global_ideal = np.min(all_solutions_array, axis=0)
            self.global_nadir = np.max(all_solutions_array, axis=0)
            self.normalization_ranges = self.global_nadir - self.global_ideal
            self.normalization_ranges[self.normalization_ranges == 0] = 1.0
           
        else:
            self.global_ideal = np.array([0.0, 0.0, 0.0])
            self.global_nadir = np.array([1.0, 1.0, 1.0])
            self.normalization_ranges = np.array([1.0, 1.0, 1.0])
    
    def normalize_solutions(self, solutions):
        """Normalize solution set to [0,1] range"""
        if self.global_ideal is None:
            self.calculate_global_reference_points()
        
        solutions_array = np.array(solutions)
        validate_solutions(solutions_array, "Solution set to be normalized")
        
        if solutions_array.shape[1] != len(self.global_ideal):
            raise ValueError(f"Solution set dimension ({solutions_array.shape[1]}) does not match global reference point dimension ({len(self.global_ideal)})")
        
        normalized = (solutions_array - self.global_ideal) / self.normalization_ranges
        normalized_clipped = np.clip(normalized, 0.0, 1.0)
        
        return normalized_clipped
    
    def _calculate_pareto_mask(self, solutions):
        """Calculate Pareto dominance relations"""
        validate_solutions(solutions, "Pareto dominance computation solution set")
        
        n = len(solutions)
        if n == 0:
            return np.array([], dtype=bool)
        
        mask = np.ones(n, dtype=bool)
        for i in range(n):
            if not mask[i]:
                continue
            dominated = np.any([is_dominated(solutions[j], solutions[i]) 
                               for j in range(n) if i != j and mask[j]])
            if dominated:
                mask[i] = False
        
        return mask
    
    def calculate_global_pareto_front_per_job(self):
        """Calculate global Pareto front for each job"""
        self.calculate_global_reference_points()
        self.job_pareto_fronts = {}
        max_jobs = self._get_max_jobs()
        
        for job_idx in range(max_jobs):
            all_solutions, solution_info = self._collect_job_solutions(job_idx)
            
            if all_solutions:
                try:
                    solutions_array = np.array(all_solutions)
                    validate_solutions(solutions_array, f"Job {job_idx} solution set")
                    
                    pareto_mask = self._calculate_pareto_mask(solutions_array)
                    
                    self.job_pareto_fronts[job_idx] = {
                        'global_pareto_solutions': solutions_array[pareto_mask],
                        'global_pareto_info': [info for i, info in enumerate(solution_info) if pareto_mask[i]],
                        'all_solutions': solutions_array,
                        'all_info': solution_info
                    }
                except Exception as e:
                    self.job_pareto_fronts[job_idx] = {
                        'global_pareto_solutions': np.array([]),
                        'global_pareto_info': [],
                        'all_solutions': np.array([]),
                        'all_info': []
                    }
        
        self._calculate_algorithm_non_dominated_cache()
    
    def _calculate_algorithm_non_dominated_cache(self):
        """Calculate and cache non-dominated solution info for each algorithm"""
        self.rule_non_dominated_cache = {}
        self.algorithm_pareto_solutions = {}
        
        for rule_name in self.all_experiment_data.keys():
            all_solutions = []
            all_info = []
            
            for job_idx, job_data in self.job_pareto_fronts.items():
                for i, info in enumerate(job_data['all_info']):
                    if info['rule_name'] == rule_name:
                        all_solutions.append(job_data['all_solutions'][i])
                        all_info.append(info)
            
            if all_solutions:
                try:
                    solutions_array = np.array(all_solutions)
                    validate_solutions(solutions_array, f"Algorithm {rule_name} solution set")
                    
                    pareto_mask = self._calculate_pareto_mask(solutions_array)
                    pareto_solutions = solutions_array[pareto_mask]
                    normalized = self.normalize_solutions(pareto_solutions)
                    
                    self.algorithm_pareto_solutions[rule_name] = normalized
                    self.rule_non_dominated_cache[rule_name] = {
                        'total_count': len(normalized),
                        'unique_solutions': normalized,
                        'all_occurrences': all_info
                    }
                except Exception as e:
                    self.algorithm_pareto_solutions[rule_name] = np.array([]).reshape(0, len(self.global_ideal))
                    self.rule_non_dominated_cache[rule_name] = {
                        'total_count': 0,
                        'unique_solutions': [],
                        'all_occurrences': []
                    }
            else:
                self.algorithm_pareto_solutions[rule_name] = np.array([]).reshape(0, len(self.global_ideal))
                self.rule_non_dominated_cache[rule_name] = {
                    'total_count': 0,
                    'unique_solutions': [],
                    'all_occurrences': []
                }
        
        # Compute nominal global Pareto front
        all_pareto_solutions = []
        for solutions in self.algorithm_pareto_solutions.values():
            if len(solutions) > 0:
                all_pareto_solutions.extend(solutions)
        
        if all_pareto_solutions:
            try:
                all_pareto_array = np.array(all_pareto_solutions)
                validate_solutions(all_pareto_array, "Global Pareto solution set")
                
                global_mask = self._calculate_pareto_mask(all_pareto_array)
                self.global_pareto_front = all_pareto_array[global_mask]
            except Exception as e:
                self.global_pareto_front = np.array([]).reshape(0, len(self.global_ideal))
        else:
            self.global_pareto_front = np.array([]).reshape(0, len(self.global_ideal))
    
    def _get_max_jobs(self):
        """Get maximum number of jobs"""
        max_jobs = 0
        for rule_data in self.all_experiment_data.values():
            for run_data in rule_data.values():
                max_jobs = max(max_jobs, len(run_data))
        return max_jobs
    
    def _collect_job_solutions(self, job_idx):
        """Collect all solutions for a job"""
        all_solutions, solution_info = [], []
        for rule_name, rule_data in self.all_experiment_data.items():
            for run_id, run_data in rule_data.items():
                if job_idx < len(run_data) and 'objectives' in run_data[job_idx]:
                    all_solutions.append(run_data[job_idx]['objectives'])
                    solution_info.append({
                        'rule_name': rule_name,
                        'run_id': run_id,
                        'objectives': run_data[job_idx]['objectives']
                    })
        return all_solutions, solution_info
    
    # ========== Three Core Metric Computation Methods ==========
    def _calculate_hypervolume(self, solutions: np.ndarray, ref_point: Optional[np.ndarray] = None) -> float:
        """Calculate Hypervolume (HV) - higher is better"""
        if len(solutions) == 0:
            return 0.0
        
        validate_solutions(solutions, "HV computation solution set")
        n_dim = solutions.shape[1]
        
        if ref_point is None:
            ref_point = np.ones(n_dim) * 1.01
        else:
            if len(ref_point) != n_dim:
                raise ValueError(f"Reference point dimension ({len(ref_point)}) does not match solution set dimension ({n_dim})")
        from pymoo.indicators.hv import HV
        hv_indicator = HV(ref_point=ref_point)
        hv_value = hv_indicator(solutions)
        return float(hv_value)
    
    def _calculate_igd(self, solutions: np.ndarray, reference_front: np.ndarray) -> float:
        """Calculate Inverted Generational Distance (IGD) - lower is better"""
        if len(solutions) == 0 or len(reference_front) == 0:
            return 1.0
        
        validate_solutions(solutions, "IGD computation solution set")
        validate_solutions(reference_front, "IGD reference front")
        
        if solutions.shape[1] != reference_front.shape[1]:
            raise ValueError(f"Solution set dimension ({solutions.shape[1]}) does not match reference front dimension ({reference_front.shape[1]})")
        
        tree = KDTree(solutions)
        distances, _ = tree.query(reference_front, k=1)
        igd_value = np.mean(distances)
        return float(igd_value)
    
    def _calculate_spread(self, solutions: np.ndarray) -> float:
        """Calculate Spread metric (Δ) - Deb standard formula, lower is better"""
        n = len(solutions)
        if n <= 1:
            return 1.0
        
        n_dim = solutions.shape[1]
        
        min_solutions = []
        max_solutions = []
        for dim in range(n_dim):
            min_idx = np.argmin(solutions[:, dim])
            max_idx = np.argmax(solutions[:, dim])
            min_solutions.append(solutions[min_idx])
            max_solutions.append(solutions[max_idx])
        
        d_f = np.linalg.norm(min_solutions[0] - max_solutions[0])
        d_l = np.linalg.norm(min_solutions[-1] - max_solutions[-1])
        d_extreme = d_f + d_l
        
        distances = calculate_nearest_neighbor_distances(solutions)
        mean_dist = np.mean(distances)
        
        numerator = d_extreme + np.sum(np.abs(distances - mean_dist))
        denominator = d_extreme + (n - 1) * mean_dist
        
        if denominator == 0:
            return 1.0
        
        return float(numerator / denominator)
    
    # ========== Main Interface Methods ==========
    def calculate_rule_performance_metrics(self) -> Dict[str, Dict]:
        """Calculate HV, IGD, and Spread metrics for each algorithm"""
        if not self.job_pareto_fronts:
            self.calculate_global_pareto_front_per_job()
        
        rule_metrics = {}
        
        for rule_name in self.all_experiment_data.keys():
            solutions = self.algorithm_pareto_solutions.get(rule_name, np.array([]).reshape(0, len(self.global_ideal)))
            
            if len(solutions) > 0:
                hypervolume = self._calculate_hypervolume(solutions)
                igd = self._calculate_igd(solutions, self.global_pareto_front)
                spread = self._calculate_spread(solutions)
                non_dominated_count = len(solutions)
            else:
                hypervolume = 0.0
                igd = 1.0
                spread = 1.0
                non_dominated_count = 0
            
            rule_metrics[rule_name] = {
                'hypervolume': hypervolume,
                'igd': igd,
                'spread': spread,
                'non_dominated_count': non_dominated_count,
                'total_jobs': len(self.job_pareto_fronts),
            }
            
        
        return rule_metrics
    
    def generate_comprehensive_report(self) -> Dict:
        """Generate comprehensive analysis report (HV, IGD, Spread only)"""
        
        rule_metrics = self.calculate_rule_performance_metrics()
        
        return {
            'rule_metrics': rule_metrics,
            'total_jobs': len(self.job_pareto_fronts),
            'total_rules': len(self.all_experiment_data),
            'global_ideal': self.global_ideal,
            'global_nadir': self.global_nadir,
            'normalization_ranges': self.normalization_ranges,
        }
    
    def save_to_excel(self, report: Dict, scenario_id: str, cyc: int,  benchmark: Optional[List[str]] = None, run_time=None, cpath: str = 'ablation') -> None:
        """Save HV, IGD, and Spread metrics to Excel"""
        if benchmark is None:
            benchmark = list(self.all_experiment_data.keys())
        
        metric_config = {
            'HV': {
                'sheet_name': 'Hypervolume(HV)',
                'metric_key': 'hypervolume',
                'header_name': 'Hypervolume',
                'evaluation_direction': 'Higher is better'
            },
            'IGD': {
                'sheet_name': 'Inverted Generational Distance',
                'metric_key': 'igd',
                'header_name': 'IGD Metric',
                'evaluation_direction': 'Lower is better'
            },
            'Spread': {
                'sheet_name': 'Spread(Diversity)',
                'metric_key': 'spread',
                'header_name': 'Spread Metric',
                'evaluation_direction': 'Lower is better'
            }
        }
        
        excel_path = Path(sys.path[0]) / "test_result" / cpath / f"{scenario_id}_Metrics_Detail.xlsx"
        
        try:
            excel_path.parent.mkdir(parents=True, exist_ok=True)
            rule_metrics = report['rule_metrics']
            
            if excel_path.exists():
                wb = load_workbook(excel_path)
            else:
                wb = Workbook()
                if 'Sheet' in wb.sheetnames:
                    del wb['Sheet']
            
            for metric_type, config in metric_config.items():
                sheet_name = config['sheet_name']
                metric_key = config['metric_key']
                
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                else:
                    ws = wb.create_sheet(title=sheet_name)
                    headers = ['Scenario', 'Run'] + benchmark
                    for col_idx, header in enumerate(headers, 1):
                        ws.cell(row=1, column=col_idx, value=header)
                
                new_row = [scenario_id, cyc + 1]
                for algo in benchmark:
                    if algo in rule_metrics:
                        metric_value = rule_metrics[algo].get(metric_key, 0.0)
                        new_row.append(metric_value)
                    else:
                        new_row.append(0.0)
                
                next_row = ws.max_row + 1
                for col_idx, value in enumerate(new_row, 1):
                    ws.cell(row=next_row, column=col_idx, value=value)
            # ========== 2. Handle runtime worksheet (only creates when run_time is not None) ==========
            if run_time is not None:
                sheet_name = "Runtime"
                # Check if runtime worksheet already exists
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                else:
                    ws = wb.create_sheet(title=sheet_name)
                    # Header format consistent with other worksheets
                    headers = ['Scenario', 'Run'] + benchmark
                    for col_idx, header in enumerate(headers, 1):
                        ws.cell(row=1, column=col_idx, value=header)
                
                # Construct runtime row data
                new_row = [scenario_id, cyc + 1]
                for algo in benchmark:
                    # Get runtime for corresponding algorithm, fill 0 if no data
                    value = run_time.get(algo, 0)
                    new_row.append(value)
                
                # Write data to next row
                next_row = ws.max_row + 1
                for col_idx, value in enumerate(new_row, 1):
                    ws.cell(row=next_row, column=col_idx, value=value)
            
            wb.save(excel_path)
            wb.close()
        
        except Exception as e:
            raise


# ============================================================================
# Part 2: Experiment Data Analysis and Visualization
# ============================================================================

class ExperimentDataLoader:
    """Experiment data loader - loads HV, IGD, and Spread metrics"""
    
    METRIC_CONFIG = {
        'HV': {
            'sheet_name': 'Hypervolume(HV)',
            'short_name': 'HV',
            'direction': 'higher',
            'description': 'Hypervolume metric (higher is better)'
        },
        'IGD': {
            'sheet_name': 'Inverted Generational Distance',
            'short_name': 'IGD',
            'direction': 'lower',
            'description': 'IGD metric (lower is better)'
        },
        'Spread': {
            'sheet_name': 'Spread(Diversity)',
            'short_name': 'Spread',
            'direction': 'lower',
            'description': 'Spread metric (lower is better)'
        }
    }
    
    def __init__(self, data_path: str = 'test_result'):
        if len(sys.path) > 0:
            project_root = sys.path[0]
        else:
            project_root = os.getcwd()
        
        self.data_path = Path(os.path.join(project_root, data_path))
        self.raw_data = {}
        self.processed_data = None
        self.summary_stats = {}
        self.benchmark_algorithms = None
    
    def get_metric_config(self):
        return self.METRIC_CONFIG.copy()
    
    def get_metric_names(self):
        return [config['short_name'] for config in self.METRIC_CONFIG.values()]
    
    def get_metric_direction(self, metric_name: str) -> str:
        for config in self.METRIC_CONFIG.values():
            if config['short_name'] == metric_name:
                return config['direction']
        return 'higher'
    
    def load_experiment_file(self, file_path: Union[str, Path]) -> Dict:
        file_path = Path(file_path)
        scenario_id = file_path.stem.replace('_Metrics_Detail', '')
        
        result = {
            'scenario': scenario_id,
            'file_path': file_path,
            'metrics': {}
        }
        
        try:
            wb = load_workbook(file_path, data_only=True)
            
            # Note: use config['sheet_name'] rather than the key name
            for metric_key, config in self.METRIC_CONFIG.items():
                sheet_name = config['sheet_name']  # ← Key: use actual worksheet name
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    
                    # Read headers
                    headers = []
                    for col in range(1, ws.max_column + 1):
                        cell_value = ws.cell(row=1, column=col).value
                        if cell_value:
                            headers.append(cell_value)
                    
                    if self.benchmark_algorithms is None and len(headers) > 2:
                        self.benchmark_algorithms = headers[2:]
                    
                    data_rows = []
                    for row in range(2, ws.max_row + 1):
                        row_data = {}
                        row_has_data = False
                        
                        for col, header in enumerate(headers, 1):
                            cell_value = ws.cell(row=row, column=col).value
                            if cell_value is not None:
                                row_data[header] = cell_value
                                if col > 2:
                                    row_has_data = True
                        
                        if row_has_data:
                            data_rows.append(row_data)
                    
                    if data_rows:
                        df = pd.DataFrame(data_rows)
                        for col in df.columns:
                            if col not in ['Scenario', 'Run']:
                                df[col] = pd.to_numeric(df[col], errors='coerce')
                        
                        # Use metric_key as storage key ('HV', 'IGD', 'Spread')
                        result['metrics'][metric_key] = {
                            'data': df,
                            'direction': config['direction'],
                            'sheet_name': sheet_name,
                            'description': config['description']
                        }
            
            wb.close()
            
        except Exception as e:
            print(f"  Error loading file {file_path.name}: {e}")
        
        return result
  
    def load_all_experiments(self, sub_dir: str = 'ablation') -> Dict:
        target_dir = self.data_path / sub_dir
        
        if not target_dir.exists():
            target_dir = self.data_path
        
        excel_files = list(target_dir.glob("*Metrics_Detail.xlsx"))
        
        if not excel_files:
            print(f"Error: No metrics detail files found in {target_dir}")
            return {}
        
        print(f"Found {len(excel_files)} experiment files")
        
        for file_path in excel_files:
            print(f"Loading file: {file_path.name}")
            experiment_data = self.load_experiment_file(file_path)
            if experiment_data['metrics']:
                scenario = experiment_data['scenario']
                self.raw_data[scenario] = experiment_data
            else:
                print(f"  Warning: No valid data in file {file_path.name}")
        
        if self.raw_data:
            self._process_data()
            metric_count = len(self.get_metric_names())
            print(f"Successfully loaded data from {len(self.raw_data)} scenarios, with {metric_count} metrics total")
        else:
            print("Error: No data was loaded successfully")
        
        return self.raw_data
    
    def _process_data(self):
        all_data = []
        
        for scenario, scenario_data in self.raw_data.items():
            for metric_name, metric_info in scenario_data['metrics'].items():
                df = metric_info['data'].copy()
                
                for algo in self.benchmark_algorithms:
                    if algo in df.columns:
                        values = df[algo].dropna().values
                        for i, value in enumerate(values):
                            all_data.append({
                                'Scenario': scenario,
                                'Algorithm': algo,
                                'Metric': metric_name,
                                'Value': value,
                                'Run': i + 1,
                                'Direction': metric_info['direction']
                            })
        
        if all_data:
            self.processed_data = pd.DataFrame(all_data)
            self._calculate_summary_stats()
            print(f"Processing complete, {len(self.processed_data)} data records in total")
    
    def _calculate_summary_stats(self):
        if self.processed_data is None or len(self.processed_data) == 0:
            return
        
        self.summary_stats['mean'] = self.processed_data.groupby(
            ['Scenario', 'Algorithm', 'Metric']
        )['Value'].mean().reset_index()
        
        self.summary_stats['std'] = self.processed_data.groupby(
            ['Scenario', 'Algorithm', 'Metric']
        )['Value'].std().reset_index()
        
        self.summary_stats['overall_mean'] = self.processed_data.groupby(
            ['Algorithm', 'Metric']
        )['Value'].mean().reset_index()
        
        self.summary_stats['overall_std'] = self.processed_data.groupby(
            ['Algorithm', 'Metric']
        )['Value'].std().reset_index()
        
        self.summary_stats['overall_median'] = self.processed_data.groupby(
            ['Algorithm', 'Metric']
        )['Value'].median().reset_index()
        
        self.summary_stats['overall_min'] = self.processed_data.groupby(
            ['Algorithm', 'Metric']
        )['Value'].min().reset_index()
        
        self.summary_stats['overall_max'] = self.processed_data.groupby(
            ['Algorithm', 'Metric']
        )['Value'].max().reset_index()
    
    def get_data_for_metric(self, metric_name: str) -> pd.DataFrame:
        if self.processed_data is None:
            return pd.DataFrame()
        return self.processed_data[self.processed_data['Metric'] == metric_name].copy()
    
    def get_algorithm_ranking(self) -> pd.DataFrame:
        """Calculate comprehensive algorithm ranking (based on normalized scores of three metrics)"""
        if self.summary_stats is None or 'overall_mean' not in self.summary_stats:
            return pd.DataFrame()
        
        overall_mean = self.summary_stats['overall_mean'].copy()
        
        normalized_values = []
        for metric in overall_mean['Metric'].unique():
            metric_df = overall_mean[overall_mean['Metric'] == metric].copy()
            direction = self.get_metric_direction(metric)
            
            min_val = metric_df['Value'].min()
            max_val = metric_df['Value'].max()
            
            if max_val == min_val:
                metric_df['Normalized Score'] = 0.5
            elif direction == 'higher':
                metric_df['Normalized Score'] = (metric_df['Value'] - min_val) / (max_val - min_val)
            else:
                metric_df['Normalized Score'] = 1 - (metric_df['Value'] - min_val) / (max_val - min_val)
            
            normalized_values.append(metric_df)
        
        if normalized_values:
            normalized_df = pd.concat(normalized_values, ignore_index=True)
            ranking = normalized_df.groupby('Algorithm')['Normalized Score'].mean().reset_index()
            ranking = ranking.sort_values('Normalized Score', ascending=False)
            ranking['Rank'] = range(1, len(ranking) + 1)
            ranking['Normalized Score'] = ranking['Normalized Score'].round(4)
            return ranking
        
        return pd.DataFrame()
    
    def get_algorithm_stats(self) -> pd.DataFrame:
        """Get detailed statistical info for algorithms"""
        if self.summary_stats is None:
            return pd.DataFrame()
        
        stats_list = []
        for algo in self.benchmark_algorithms:
            for metric in self.get_metric_names():
                mean_df = self.summary_stats['overall_mean']
                std_df = self.summary_stats['overall_std']
                median_df = self.summary_stats['overall_median']
                min_df = self.summary_stats['overall_min']
                max_df = self.summary_stats['overall_max']
                
                mean_val = mean_df[(mean_df['Algorithm'] == algo) & (mean_df['Metric'] == metric)]['Value'].values
                std_val = std_df[(std_df['Algorithm'] == algo) & (std_df['Metric'] == metric)]['Value'].values
                median_val = median_df[(median_df['Algorithm'] == algo) & (median_df['Metric'] == metric)]['Value'].values
                min_val = min_df[(min_df['Algorithm'] == algo) & (min_df['Metric'] == metric)]['Value'].values
                max_val = max_df[(max_df['Algorithm'] == algo) & (max_df['Metric'] == metric)]['Value'].values
                
                if len(mean_val) > 0:
                    stats_list.append({
                        'Algorithm': algo,
                        'Metric': metric,
                        'Mean': mean_val[0],
                        'Std': std_val[0] if len(std_val) > 0 else np.nan,
                        'Median': median_val[0] if len(median_val) > 0 else np.nan,
                        'Min': min_val[0] if len(min_val) > 0 else np.nan,
                        'Max': max_val[0] if len(max_val) > 0 else np.nan
                    })
        
        return pd.DataFrame(stats_list)
    
    def generate_detailed_scenario_metrics(self, output_path: Union[str, Path], 
                                            significance_levels: Dict = None):
        """Generate fine-grained scenario metrics Excel file (HV, IGD, Spread)"""
        if self.processed_data is None or len(self.processed_data) == 0:
            return
        
        if significance_levels is None:
            significance_levels = {'**': 0.01, '*': 0.05, '†': 0.10}
        
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        scenarios = sorted(self.processed_data['Scenario'].unique())
        algorithms = self.benchmark_algorithms
        
        metrics_info = {}
        for metric in self.get_metric_names():
            metrics_info[metric] = {
                'sheet_name': metric,
                'direction': self.get_metric_direction(metric),
                'description': self.METRIC_CONFIG[metric]['description']
            }
        
        all_sheets = {}
        
        for metric_name, metric_cfg in metrics_info.items():
            metric_data = self.processed_data[self.processed_data['Metric'] == metric_name]
            
            if metric_data.empty:
                continue
            
            pivot_mean = metric_data.groupby(['Scenario', 'Algorithm'])['Value'].mean().unstack()
            pivot_std = metric_data.groupby(['Scenario', 'Algorithm'])['Value'].std().unstack()
            
            for algo in algorithms:
                if algo not in pivot_mean.columns:
                    pivot_mean[algo] = np.nan
                    pivot_std[algo] = np.nan
            
            pivot_mean = pivot_mean[algorithms]
            pivot_std = pivot_std[algorithms]
            
            display_df = pd.DataFrame(index=pivot_mean.index, columns=pivot_mean.columns)
            best_flags = pd.DataFrame(index=pivot_mean.index, columns=pivot_mean.columns, dtype=bool)
            
            for scenario in pivot_mean.index:
                scenario_means = pivot_mean.loc[scenario].copy()
                scenario_means = scenario_means[~pd.isna(scenario_means)]
                
                if len(scenario_means) == 0:
                    for algo in algorithms:
                        display_df.loc[scenario, algo] = 'N/A'
                        best_flags.loc[scenario, algo] = False
                    continue
                
                if metric_cfg['direction'] == 'higher':
                    best_algo = scenario_means.idxmax()
                else:
                    best_algo = scenario_means.idxmin()
                
                for algo in algorithms:
                    mean_val = pivot_mean.loc[scenario, algo]
                    std_val = pivot_std.loc[scenario, algo]
                    
                    if pd.isna(mean_val) or pd.isna(std_val):
                        display_df.loc[scenario, algo] = 'N/A'
                        best_flags.loc[scenario, algo] = False
                        continue
                    
                    base_str = f"{mean_val:.4f} ± {std_val:.4f}"
                    is_best = (algo == best_algo)
                    best_flags.loc[scenario, algo] = is_best
                    
                    superscript = ''
                    if not is_best:
                        try:
                            algo_data = metric_data[
                                (metric_data['Scenario'] == scenario) & 
                                (metric_data['Algorithm'] == algo)
                            ]['Value'].values
                            
                            best_data = metric_data[
                                (metric_data['Scenario'] == scenario) & 
                                (metric_data['Algorithm'] == best_algo)
                            ]['Value'].values
                            
                            if len(algo_data) > 1 and len(best_data) > 1:
                                _, p_value = stats.ttest_ind(algo_data, best_data, equal_var=False)
                                
                                for sym, level in sorted(significance_levels.items(), key=lambda x: x[1]):
                                    if p_value < level:
                                        superscript = sym
                                        break
                        except Exception as e:
                            pass
                    
                    display_df.loc[scenario, algo] = base_str + superscript
            
            all_sheets[metric_cfg['sheet_name']] = {
                'data': display_df,
                'best_flags': best_flags,
                'direction': metric_cfg['direction'],
                'description': metric_cfg['description']
            }
        
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            from openpyxl.styles import Font, Alignment
            
            for sheet_name, sheet_content in all_sheets.items():
                display_df = sheet_content['data']
                best_flags = sheet_content['best_flags']
                description = sheet_content.get('description', '')
                
                display_df.to_excel(writer, sheet_name=sheet_name)
                
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]
                
                center_align = Alignment(horizontal='center', vertical='center')
                bold_font = Font(bold=True)
                
                for col_idx, algo in enumerate(display_df.columns, start=2):
                    cell = worksheet.cell(row=1, column=col_idx)
                    cell.font = bold_font
                    cell.alignment = center_align
                    cell.value = algo
                
                for row_idx, scenario in enumerate(display_df.index, start=2):
                    scene_cell = worksheet.cell(row=row_idx, column=1)
                    scene_cell.value = str(scenario)
                    scene_cell.alignment = center_align
                    
                    for col_idx, algo in enumerate(display_df.columns, start=2):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.alignment = center_align
                        
                        if not best_flags.empty and scenario in best_flags.index and algo in best_flags.columns:
                            if best_flags.loc[scenario, algo]:
                                cell.font = bold_font
    
        
        return str(output_path)
    
    def generate_win_rate_statistics(self) -> pd.DataFrame:
        """Generate algorithm win rate statistics table (based on three metrics)"""
        if self.processed_data is None or len(self.processed_data) == 0:
            return pd.DataFrame()
        
        metrics = self.get_metric_names()
        scenarios = self.processed_data['Scenario'].unique()
        algorithms = self.benchmark_algorithms
        
        stats_dict = {algo: {'Best Count': 0, 'Sig Better Count': 0, 'Sig Worse Count': 0} for algo in algorithms}
        
        for metric in metrics:
            metric_data = self.processed_data[self.processed_data['Metric'] == metric]
            direction = self.get_metric_direction(metric)
            
            for scenario in scenarios:
                scenario_data = metric_data[metric_data['Scenario'] == scenario]
                
                if scenario_data.empty:
                    continue
                
                algo_means = {}
                algo_values = {}
                
                for algo in algorithms:
                    algo_data = scenario_data[scenario_data['Algorithm'] == algo]['Value'].values
                    if len(algo_data) > 0:
                        algo_means[algo] = np.mean(algo_data)
                        algo_values[algo] = algo_data
                
                if not algo_means:
                    continue
                
                if direction == 'higher':
                    best_algo = max(algo_means, key=algo_means.get)
                else:
                    best_algo = min(algo_means, key=algo_means.get)
                
                stats_dict[best_algo]['Best Count'] += 1
                
                for algo in algorithms:
                    if algo == best_algo or algo not in algo_values:
                        continue
                    
                    algo_data = algo_values[algo]
                    best_data = algo_values[best_algo]
                    
                    if len(algo_data) >= 2 and len(best_data) >= 2:
                        try:
                            _, p_value = stats.ttest_ind(algo_data, best_data, equal_var=False)
                            if p_value < 0.05:
                                stats_dict[algo]['Sig Worse Count'] += 1
                                stats_dict[best_algo]['Sig Better Count'] += 1
                        except:
                            pass
        
        total_scenarios = len(scenarios) * len(metrics)
        result = []
        for algo in algorithms:
            s = stats_dict[algo]
            best_ratio = s['Best Count'] / total_scenarios if total_scenarios > 0 else 0
            sig_better_ratio = s['Sig Better Count'] / total_scenarios if total_scenarios > 0 else 0
            sig_worse_ratio = s['Sig Worse Count'] / total_scenarios if total_scenarios > 0 else 0
            comprehensive_win_rate = (s['Best Count'] + 0.5 * s['Sig Better Count']) / total_scenarios if total_scenarios > 0 else 0
            
            result.append({
                'Algorithm': algo,
                'Best Count': s['Best Count'],
                'Best Ratio': f"{best_ratio:.2%}",
                'Sig Better Count': s['Sig Better Count'],
                'Sig Better Ratio': f"{sig_better_ratio:.2%}",
                'Sig Worse Count': s['Sig Worse Count'],
                'Sig Worse Ratio': f"{sig_worse_ratio:.2%}",
                'Combined Win Rate': f"{comprehensive_win_rate:.2%}",
                'Combined Score': comprehensive_win_rate
            })
        
        result_df = pd.DataFrame(result)
        result_df = result_df.sort_values('Combined Score', ascending=False)
        result_df = result_df.drop('Combined Score', axis=1)
        result_df['Rank'] = range(1, len(result_df) + 1)
        
        column_order = ['Rank', 'Algorithm', 'Best Count', 'Best Ratio', 'Sig Better Count', 'Sig Better Ratio', 
                        'Sig Worse Count', 'Sig Worse Ratio', 'Combined Win Rate']
        result_df = result_df[column_order]
        
        return result_df


class ExperimentAnalyzer:
    """Experiment data analyzer - supports HV, IGD, and Spread metrics"""
    
    def __init__(self, data_loader: ExperimentDataLoader):
        self.data_loader = data_loader
        self.processed_data = data_loader.processed_data
        self.summary_stats = data_loader.summary_stats
        self.benchmark_algorithms = data_loader.benchmark_algorithms
    
    def friedman_test(self, metric_name: str) -> Dict:
        """Friedman test"""
        df = self.data_loader.get_data_for_metric(metric_name)
        
        if df.empty:
            return {'error': 'No data', 'metric': metric_name}
        
        pivot_df = df.groupby(['Scenario', 'Algorithm'])['Value'].mean().reset_index()
        pivot_df = pivot_df.pivot(index='Scenario', columns='Algorithm', values='Value')
        pivot_df = pivot_df.dropna(axis=1)
        
        if pivot_df.shape[1] < 2:
            return {'error': 'Insufficient number of algorithms', 'metric': metric_name}
        if pivot_df.shape[0] < 2:
            return {'error': 'Insufficient number of scenarios', 'metric': metric_name}
        
        algorithms = pivot_df.columns.tolist()
        data_matrix = [pivot_df[algo].values for algo in algorithms]
        
        try:
            stat, p_value = friedmanchisquare(*data_matrix)
            return {
                'metric': metric_name,
                'statistic': stat,
                'p_value': p_value,
                'significant': p_value < 0.05,
                'algorithms': algorithms,
                'data_matrix': pivot_df
            }
        except Exception as e:
            return {'error': str(e), 'metric': metric_name}
    
    def nemenyi_posthoc(self, metric_name: str) -> pd.DataFrame:
        """Nemenyi post-hoc test"""
        friedman_result = self.friedman_test(metric_name)
        
        if 'error' in friedman_result:
            return pd.DataFrame()
        
        pivot_df = friedman_result['data_matrix'].copy()
        
        try:
            pivot_df = pivot_df.dropna(axis=1, how='any')
            pivot_df = pivot_df.dropna(axis=0, how='any')
            
            if pivot_df.shape[1] < 2:
                return pd.DataFrame()
            
            algorithms = pivot_df.columns.tolist()
            data_for_test = pivot_df.values.T
            nemenyi_result = sp.posthoc_nemenyi_friedman(data_for_test)
            
            result_df = pd.DataFrame(
                nemenyi_result,
                index=algorithms[:nemenyi_result.shape[0]],
                columns=algorithms[:nemenyi_result.shape[1]]
            )
            return result_df
        except Exception as e:
            return pd.DataFrame()
    
    def calculate_statistical_summary(self) -> Dict:
        """Calculate complete statistical summary"""
        summary = {
            'friedman_tests': {},
            'ranking': self.data_loader.get_algorithm_ranking(),
        }
        
        for metric in self.data_loader.get_metric_names():
            friedman_result = self.friedman_test(metric)
            summary['friedman_tests'][metric] = friedman_result
            
            if 'error' not in friedman_result and friedman_result['significant']:
                nemenyi_result = self.nemenyi_posthoc(metric)
                summary['friedman_tests'][metric]['nemenyi'] = nemenyi_result
        
        return summary


class ExperimentVisualizer:
    """Experiment data visualizer - supports HV, IGD, and Spread metrics"""
    
    def __init__(self, data_loader: ExperimentDataLoader, analyzer: ExperimentAnalyzer):
        self.data_loader = data_loader
        self.analyzer = analyzer
    
    def plot_boxplots(self, save_path: Optional[str] = None):
        """Plot boxplots for each metric"""
        metrics = self.data_loader.get_metric_names()
        n_metrics = len(metrics)
        
        fig, axes = plt.subplots(1, n_metrics, figsize=(6*n_metrics, 6))
        if n_metrics == 1:
            axes = [axes]
        
        ranking = self.data_loader.get_algorithm_ranking()
        
        for idx, metric in enumerate(metrics):
            ax = axes[idx]
            df = self.data_loader.get_data_for_metric(metric)
            
            if df.empty:
                ax.text(0.5, 0.5, 'No data', ha='center', va='center')
                ax.set_title(f'{metric}')
                continue
            
            direction = self.data_loader.get_metric_direction(metric)
            algo_means = df.groupby('Algorithm')['Value'].mean()
            
            if direction == 'higher':
                algo_order = algo_means.sort_values(ascending=False).index.tolist()
            else:
                algo_order = algo_means.sort_values(ascending=True).index.tolist()
            
            colors = sns.color_palette("viridis", n_colors=len(algo_order))
            palette = {algo: colors[i] for i, algo in enumerate(algo_order)}
            
            sns.boxplot(data=df, x='Algorithm', y='Value', ax=ax, order=algo_order, palette=palette)
            
            ax.set_title(f'{metric} Metric Comparison', fontsize=12, fontweight='bold')
            ax.set_xlabel('')
            ax.set_ylabel('Value')
            ax.tick_params(axis='x', rotation=45)
            ax.grid(True, alpha=0.3, axis='y')
            
            friedman_result = self.analyzer.friedman_test(metric)
            if 'error' not in friedman_result:
                p_text = f"Friedman p = {friedman_result['p_value']:.4f}"
                if friedman_result['significant']:
                    p_text += " *"
                ax.text(0.02, 0.98, p_text, transform=ax.transAxes, fontsize=10,
                       verticalalignment='top', bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.8))
        
        plt.suptitle('Overall Performance Comparison of Algorithms on HV, IGD, and Spread', fontsize=14, fontweight='bold')
        plt.tight_layout()
        
        if save_path:
            plt.savefig(save_path, dpi=300, bbox_inches='tight')
            print(f"Boxplot saved to: {save_path}")
        else:
            plt.show()
        plt.close()
    
    def plot_radar_chart(self, top_k: int = 8, save_path: Optional[str] = None):
        """Plot radar chart of algorithm comprehensive capability"""
        ranking = self.data_loader.get_algorithm_ranking()
        
        if ranking.empty:
            print("No ranking data")
            return
        
        top_algorithms = ranking.head(top_k)['Algorithm'].tolist()
        metrics = self.data_loader.get_metric_names()
        
        normalized_scores = []
        for metric in metrics:
            df = self.data_loader.get_data_for_metric(metric)
            direction = self.data_loader.get_metric_direction(metric)
            
            if not df.empty:
                algo_means = df.groupby('Algorithm')['Value'].mean()
                min_val = algo_means.min()
                max_val = algo_means.max()
                
                for algo in top_algorithms:
                    if algo in algo_means.index:
                        if max_val == min_val:
                            norm_score = 0.5
                        elif direction == 'higher':
                            norm_score = (algo_means[algo] - min_val) / (max_val - min_val)
                        else:
                            norm_score = 1 - (algo_means[algo] - min_val) / (max_val - min_val)
                    else:
                        norm_score = 0
                    
                    normalized_scores.append({'Algorithm': algo, 'Metric': metric, 'Score': norm_score})
        
        if not normalized_scores:
            print("No normalized data")
            return
        
        score_df = pd.DataFrame(normalized_scores)
        
        fig = plt.figure(figsize=(10, 8))
        ax = fig.add_subplot(111, projection='polar')
        
        angles = [n / len(metrics) * 2 * pi for n in range(len(metrics))]
        angles += angles[:1]
        
        colors = sns.color_palette("husl", n_colors=len(top_algorithms))
        
        for i, algo in enumerate(top_algorithms):
            algo_scores = score_df[score_df['Algorithm'] == algo].set_index('Metric')['Score']
            values = [algo_scores.get(metric, 0) for metric in metrics]
            values += values[:1]
            
            ax.plot(angles, values, 'o-', linewidth=2, label=algo, color=colors[i])
            ax.fill(angles, values, alpha=0.1, color=colors[i])
        
        ax.set_xticks(angles[:-1])
        ax.set_xticklabels(metrics, fontsize=10)
        ax.set_ylim(0, 1)
        ax.set_theta_offset(pi / 2)
        ax.set_theta_direction(-1)
        
        plt.legend(loc='upper right', bbox_to_anchor=(1.3, 1.1), fontsize=10)
        plt.title(f'Top-{top_k} Algorithm Comprehensive Capability Radar Chart (HV/IGD/Spread)', fontsize=14, fontweight='bold')
        
        if save_path:
            plt.savefig(save_path, dpi=300, bbox_inches='tight')
            print(f"Radar chart saved to: {save_path}")
        else:
            plt.show()
        plt.close()
    
    def plot_heatmap(self, metric_name: str, save_path: Optional[str] = None):
        """Plot heatmap"""
        df = self.data_loader.get_data_for_metric(metric_name)
        
        if df.empty:
            print(f"No {metric_name} data")
            return
        
        pivot_df = df.groupby(['Scenario', 'Algorithm'])['Value'].mean().reset_index()
        pivot_df = pivot_df.pivot(index='Scenario', columns='Algorithm', values='Value')
        
        direction = self.data_loader.get_metric_direction(metric_name)
        cmap = 'YlOrRd' if direction == 'higher' else 'YlGnBu_r'
        
        fig, ax = plt.subplots(figsize=(14, 8))
        sns.heatmap(pivot_df, annot=True, fmt='.4f', cmap=cmap, linewidths=0.5, ax=ax)
        
        ax.set_title(f'{metric_name} - Algorithm Performance Across Different Scenarios', fontsize=12, fontweight='bold')
        ax.set_xlabel('Algorithm')
        ax.set_ylabel('Scenario')
        ax.tick_params(axis='x', rotation=45)
        
        if save_path:
            plt.savefig(save_path, dpi=300, bbox_inches='tight')
            print(f"Heatmap saved to: {save_path}")
        else:
            plt.show()
        plt.close()
    
    def plot_statistical_significance(self, metric_name: str, save_path: Optional[str] = None):
        """Plot statistical significance matrix"""
        nemenyi_result = self.analyzer.nemenyi_posthoc(metric_name)
        
        if nemenyi_result.empty:
            print(f"Cannot plot significance matrix for {metric_name}")
            return
        
        fig, ax = plt.subplots(figsize=(10, 8))
        mask = np.triu(np.ones_like(nemenyi_result, dtype=bool))
        
        sns.heatmap(nemenyi_result, mask=mask, annot=True, fmt='.4f', 
                   cmap='RdYlGn_r', center=0.05, vmin=0, vmax=0.1,
                   linewidths=0.5, ax=ax, cbar_kws={'label': 'p-value'})
        
        ax.set_title(f'{metric_name} - Nemenyi Post-hoc Test Significance Matrix', fontsize=12, fontweight='bold')
        ax.set_xlabel('Algorithm')
        ax.set_ylabel('Algorithm')
        ax.tick_params(axis='x', rotation=45)
        
        ax.text(0.02, 0.98, '* p < 0.05', transform=ax.transAxes, fontsize=10,
               verticalalignment='top', bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.8))
        
        if save_path:
            plt.savefig(save_path, dpi=300, bbox_inches='tight')
            print(f"Significance matrix saved to: {save_path}")
        else:
            plt.show()
        plt.close()


class ExperimentReportGenerator:
    """Experiment report generator"""
    
    def __init__(self, data_loader: ExperimentDataLoader, 
                 analyzer: ExperimentAnalyzer,
                 visualizer: ExperimentVisualizer):
        self.data_loader = data_loader
        self.analyzer = analyzer
        self.visualizer = visualizer
    
    def generate_full_report(self, output_dir: str = 'analysis_report'):
        """Generate complete experiment analysis report"""
        project_root = sys.path[0] if len(sys.path) > 0 else os.getcwd()
        output_path = Path(os.path.join(project_root, output_dir))
        output_path.mkdir(parents=True, exist_ok=True)
        
        metrics = self.data_loader.get_metric_names()
        
        print("=" * 60)
        print("Starting complete experiment analysis report generation")
        print("=" * 60)
        print(f"Detected metrics: {', '.join(metrics)}")
        print("-" * 60)
        
        print("\n[1/5] Computing statistical summary...")
        statistical_summary = self.analyzer.calculate_statistical_summary()
        
        print("[2/5] Saving statistical summary to Excel...")
        self._save_statistical_summary(statistical_summary, output_path / 'statistical_summary.xlsx')
        
        print("[3/5] Generating boxplots and radar chart...")
        self.visualizer.plot_boxplots(save_path=output_path / '01_boxplots.png')
        self.visualizer.plot_radar_chart(top_k=8, save_path=output_path / '02_radar_chart.png')
        
        print("[4/5] Generating detailed analysis for each metric...")
        for metric_name in metrics:
            metric_dir = output_path / metric_name
            metric_dir.mkdir(exist_ok=True)
            
            self.visualizer.plot_heatmap(metric_name, save_path=metric_dir / f'{metric_name}_heatmap.png')
            self.visualizer.plot_statistical_significance(metric_name, 
                                                          save_path=metric_dir / f'{metric_name}_significance.png')
        
        print("[5/5] Generating text report and algorithm statistics table...")
        self._generate_text_report(statistical_summary, output_path / 'analysis_report.txt')
        self._save_algorithm_stats(output_path / 'algorithm_statistics.csv')
        
        print("\n" + "=" * 60)
        print(f"Report generation complete! All files saved to: {output_path}")
        print("=" * 60)
    
    def _save_statistical_summary(self, summary: Dict, file_path: Path):
        """Save statistical summary to Excel"""
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            if not summary['ranking'].empty:
                summary['ranking'].to_excel(writer, sheet_name='Algorithm Ranking', index=False)
            
            friedman_data = []
            for metric, result in summary['friedman_tests'].items():
                if 'error' not in result:
                    friedman_data.append({
                        'Metric': metric,
                        'Friedman Statistic': result['statistic'],
                        'p-value': result['p_value'],
                        'Significant': 'Yes' if result['significant'] else 'No'
                    })
            
            if friedman_data:
                pd.DataFrame(friedman_data).to_excel(writer, sheet_name='Friedman Test', index=False)
    
    def _save_algorithm_stats(self, file_path: Path):
        """Save algorithm detailed statistics"""
        stats_df = self.data_loader.get_algorithm_stats()
        if not stats_df.empty:
            stats_df.to_csv(file_path, index=False, encoding='utf-8-sig')
            print(f"  Algorithm statistics table saved to: {file_path}")
    
    def _generate_text_report(self, summary: Dict, file_path: Path):
        """Generate text format analysis report"""
        metrics = self.data_loader.get_metric_names()
        
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write("=" * 80 + "\n")
            f.write("Multi-Objective Dynamic Job Shop Scheduling - Reinforcement Learning Algorithm Experiment Analysis Report\n")
            f.write("=" * 80 + "\n\n")
            
            f.write("1. Data Overview\n")
            f.write("-" * 40 + "\n")
            f.write(f"  Number of experiment scenarios: {len(self.data_loader.raw_data)}\n")
            f.write(f"  Number of benchmark algorithms: {len(self.data_loader.benchmark_algorithms) if self.data_loader.benchmark_algorithms else 0}\n")
            f.write(f"  Number of evaluation metrics: {len(metrics)} ({', '.join(metrics)})\n\n")
            
            f.write("2. Algorithm Comprehensive Ranking\n")
            f.write("-" * 40 + "\n")
            if not summary['ranking'].empty:
                for _, row in summary['ranking'].iterrows():
                    f.write(f"  {int(row['Rank'])}. {row['Algorithm']}: {row['Normalized Score']:.4f}\n")
            else:
                f.write("  No ranking data\n")
            f.write("\n")
            
            f.write("3. Statistical Significance Test\n")
            f.write("-" * 40 + "\n")
            for metric, result in summary['friedman_tests'].items():
                if 'error' not in result:
                    f.write(f"\n  {metric}:\n")
                    f.write(f"    Friedman test: p = {result['p_value']:.4e} ")
                    f.write("(Significant difference)\n" if result['significant'] else "(No significant difference)\n")
                    
                    if result['significant'] and 'nemenyi' in result:
                        nemenyi = result['nemenyi']
                        if not nemenyi.empty:
                            f.write("    Nemenyi post-hoc significant difference pairs (p < 0.05):\n")
                            for i, algo1 in enumerate(nemenyi.columns):
                                for j, algo2 in enumerate(nemenyi.columns):
                                    if i < j and nemenyi.iloc[i, j] < 0.05:
                                        f.write(f"      - {algo1} vs {algo2}: p = {nemenyi.iloc[i, j]:.4f}\n")
            
            f.write("\n" + "=" * 80 + "\n")
            f.write("Report generation time: " + pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S") + "\n")
            f.write("=" * 80 + "\n")
    
    def generate_full_report_with_detailed_analysis(self, output_dir: str = 'analysis_report'):
        """Generate complete report with fine-grained analysis (enhanced version)"""
        self.generate_full_report(output_dir)
        
        output_path = Path(output_dir)
        
        print("\n[Enhanced Analysis 1/2] Generating fine-grained scenario metrics report...")
        detailed_excel_path = output_path / '01_detailed_scenario_metrics.xlsx'
        self.data_loader.generate_detailed_scenario_metrics(detailed_excel_path)
        
        print("[Enhanced Analysis 2/2] Generating algorithm win rate statistics table...")
        win_rate_df = self.data_loader.generate_win_rate_statistics()
        if not win_rate_df.empty:
            win_rate_path = output_path / '02_win_rate_statistics.csv'
            win_rate_df.to_csv(win_rate_path, index=False, encoding='utf-8-sig')
            print(f"  Win rate statistics table saved to: {win_rate_path}")
        
        print("\n" + "=" * 60)
        print("Enhanced analysis complete! All fine-grained reports have been generated.")
        print("=" * 60)


def data_analysis_report(excel_path='test_result', sub_path='Ref_Learning'):
    """Main function: run complete data analysis"""
    project_root = sys.path[0]
    base_output_path = os.path.join(project_root, "test_result", "experimental_analysis", sub_path)
    os.makedirs(base_output_path, exist_ok=True)
    
    detailed_metrics_path = os.path.join(base_output_path, "detailed_metrics.xlsx")
    win_rate_path = os.path.join(base_output_path, "win_rate_statistics.xlsx")
    
    print(f"Project root: {project_root}")
    print(f"Base output directory: {base_output_path}")
    
    loader, analyzer, visualizer, reporter = run_complete_analysis(
        data_path=excel_path, sub_dir=sub_path, output_dir=base_output_path
    )
    
    if loader and loader.processed_data is not None:
        print("\n" + "=" * 60)
        print("Starting enhanced analysis module...")
        print("=" * 60)
        
        loader.generate_detailed_scenario_metrics(detailed_metrics_path)
        print(f"Fine-grained report saved at: {os.path.abspath(detailed_metrics_path)}")
        
        win_rate = loader.generate_win_rate_statistics()
        if not win_rate.empty:
            win_rate.to_excel(win_rate_path, index=False)
            print(f"Win rate statistics table saved at: {os.path.abspath(win_rate_path)}")


def run_complete_analysis(data_path: str = 'test_result', sub_dir: str = 'ablation', output_dir: str = 'analysis_report'):
    """Run complete experiment data analysis pipeline"""
    print("=" * 60)
    print("Multi-Objective Dynamic Job Shop Scheduling - RL Algorithm Experiment Analysis System")
    print("=" * 60)
    
    print("\n[Phase 1] Loading experiment data...")
    loader = ExperimentDataLoader(data_path)
    loader.load_all_experiments(sub_dir=sub_dir)
    
    if not loader.raw_data:
        print("Error: No data was loaded successfully, please check the data path!")
        return None, None, None, None
    
    print("\n[Phase 2] Initializing data analyzer...")
    analyzer = ExperimentAnalyzer(loader)
    
    print("\n[Phase 3] Initializing visualization engine...")
    visualizer = ExperimentVisualizer(loader, analyzer)
    
    print("\n[Phase 4] Generating complete analysis report...")
    reporter = ExperimentReportGenerator(loader, analyzer, visualizer)
    reporter.generate_full_report(output_dir=output_dir)
    
    return loader, analyzer, visualizer, reporter


if __name__ == "__main__":
    data_analysis_report(excel_path='test_result', sub_path='Ref_Learning')